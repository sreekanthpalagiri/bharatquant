"""
Excel Export and Formatting Module for BharatQuant.

Handles:
1. Converting analytical results into a structured Pandas DataFrame.
2. Generating a professional Excel (.xlsx) file.
3. Applying color-coded conditional formatting for Bullish (Green) and Bearish (Red) signals.
4. Setting column widths and number formats (Currency, %, Multiples).
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from .config import log, OUTPUT_FILE, HEADER_LABELS, WIDTH_MAP, FMT_MAP, TODAY, COLOR_RULES, LEGEND

# Define Colors for UI Highlighting
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def write_excel(rows: list):
    """
    Export the processed stock data to an Excel file with advanced formatting.
    Applies conditional colors:
    - 🟩 Green for high RS Scores, high F-Scores, and Strong Growth.
    - 🟥 Red for poor growth, high debt, or high share pledging.
    """
    if not rows:
        log.warning("No data rows to write to Excel.")
        return

    log.info(f"Generating professional report: {OUTPUT_FILE}...")
    df = pd.DataFrame(rows)

    # Ensure columns match the configuration order
    cols_to_use = [c for c in HEADER_LABELS if c in df.columns]
    df = df[cols_to_use]

    # Helper to evaluate config string conditions
    def check_cond(val, rule_str):
        if val is None: return False
        if "or" in rule_str:
            return any(check_cond(val, r.strip()) for r in rule_str.split("or"))
        if "-" in rule_str and not rule_str.startswith("-"):
            try:
                p = rule_str.split("-")
                return float(p[0]) <= float(val) <= float(p[1])
            except: return False
        op = ''.join([c for c in rule_str if c in '<=>'])
        try:
            num = float(rule_str.replace(op, "").strip())
            if op == ">": return float(val) > num
            if op == ">=": return float(val) >= num
            if op == "<": return float(val) < num
            if op == "<=": return float(val) <= num
        except: pass
        return False

    def compute_fill_and_reason(row_dict):
        red_triggers = []
        for col, rule in COLOR_RULES.get("red", {}).items():
            if col == "Growth %":
                sg = row_dict.get("Sales Growth % (YoY)")
                pg = row_dict.get("Profit Growth % (YoY)")
                if check_cond(sg, rule):
                    red_triggers.append("Sales Growth < 0%")
                if check_cond(pg, rule):
                    red_triggers.append("Profit Growth < 0%")
            elif col in row_dict and check_cond(row_dict.get(col), rule):
                red_triggers.append(f"{col} {rule}")

        if red_triggers:
            return RED_FILL, "Red: " + " | ".join(red_triggers)

        green_hits = []
        for col, rule in COLOR_RULES.get("green", {}).items():
            if col == "Growth %":
                if check_cond(row_dict.get("Sales Growth % (YoY)"), rule) and check_cond(row_dict.get("Profit Growth % (YoY)"), rule):
                    green_hits.append("Growth %")
            elif col in row_dict and check_cond(row_dict.get(col), rule):
                green_hits.append(col)

        total_green = len(COLOR_RULES.get("green", {}))
        if len(green_hits) >= 4:
            return GREEN_FILL, f"Green ({len(green_hits)}/{total_green}): " + ", ".join(green_hits)

        yellow_triggers = []
        for col, rule in COLOR_RULES.get("yellow", {}).items():
            if col in row_dict and check_cond(row_dict.get(col), rule):
                yellow_triggers.append(f"{col} {rule}")

        if yellow_triggers:
            return YELLOW_FILL, "Yellow: " + " | ".join(yellow_triggers)

        return None, f"Neutral ({len(green_hits)}/{total_green}): " + (", ".join(green_hits) if green_hits else "no signals")

    # Pre-compute fills and reasons before writing to Excel
    fills = []
    reasons = []
    for row in df.itertuples(index=False):
        row_dict = dict(zip(df.columns, row))
        fill, reason = compute_fill_and_reason(row_dict)
        fills.append(fill)
        reasons.append(reason)

    df["Color Reason"] = reasons

    # Save to Excel
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Screener")
        ws = writer.sheets["Screener"]

        # Apply Formatting
        for row_idx, (row_fill, row) in enumerate(zip(fills, df.itertuples(index=False)), start=2):
            for col_idx, (lbl, val) in enumerate(zip(df.columns, row), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if lbl in FMT_MAP:
                    cell.number_format = FMT_MAP[lbl]
                if row_fill:
                    cell.fill = row_fill

        # 3. Final UI Adjustments (Header Freeze, Column Widths)
        ws.freeze_panes = "A2"
        for ci, lbl in enumerate(df.columns, start=1):
            width = WIDTH_MAP.get(lbl, 12) if lbl != "Color Reason" else 50
            ws.column_dimensions[get_column_letter(ci)].width = width

        # 4. Add Legend Sheet
        df_legend = pd.DataFrame(LEGEND)
        df_legend.to_excel(writer, index=False, sheet_name="Legend")
        ws_legend = writer.sheets["Legend"]
        ws_legend.column_dimensions["A"].width = 25
        ws_legend.column_dimensions["B"].width = 45
        ws_legend.column_dimensions["C"].width = 45
        ws_legend.column_dimensions["D"].width = 30

    log.info(f"Report Successfully Generated! ✨ → {os.path.abspath(OUTPUT_FILE)}")
